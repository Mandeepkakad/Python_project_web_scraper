import requests
import openpyxl
from bs4 import BeautifulSoup

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movie Name', 'Year of release', 'IMDB Rating'])

try:

    source = requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')

    movies = soup.find('tbody', class_="lister-list").find_all('tr')
    print(len(movies))

    for movie in movies:

        name = movie.find('td', class_="titleColumn").a.text

        rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]

        year = movie.find('span', class_="secondaryInfo").text.strip('()')

        rating = movie.find('td', class_="ratingColumn imdbRating").strong.text
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])


except e as Exception:
    print(e)

excel.save('IMDBS_movie_rating.xlsx')